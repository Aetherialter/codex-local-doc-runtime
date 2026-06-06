from __future__ import annotations

from pathlib import Path

from docrt.paths import ensure_unlocked_for_read, validate_input_path
from docrt.runtime_env import assert_mainline_runtime_for_path


def inspect_xlsx(path: str | Path) -> dict[str, object]:
    input_path = validate_input_path(path, {".xlsx"})
    assert_mainline_runtime_for_path(input_path)
    ensure_unlocked_for_read(input_path)
    try:
        import openpyxl
    except Exception as exc:
        raise RuntimeError("openpyxl is unavailable") from exc

    workbook = openpyxl.load_workbook(str(input_path), data_only=False, read_only=False)
    try:
        sheets = []
        for sheet in workbook.worksheets:
            preview = []
            max_row = min(sheet.max_row or 0, 20)
            max_col = min(sheet.max_column or 0, 20)
            for row in sheet.iter_rows(
                min_row=1,
                max_row=max_row,
                min_col=1,
                max_col=max_col,
                values_only=True,
            ):
                preview.append(list(row))
            sheets.append(
                {
                    "name": sheet.title,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                    "merged_cells": [str(rng) for rng in sheet.merged_cells.ranges],
                    "preview": preview,
                }
            )
        return {"path": str(input_path), "sheet_count": len(sheets), "sheets": sheets}
    finally:
        workbook.close()
