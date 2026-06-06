from __future__ import annotations

import shutil
from pathlib import Path
from typing import Any

from docrt.models import ErrorCode
from docrt.patch_common import (
    ensure_distinct_output,
    load_patch,
    patch_summary,
    require_string,
    verification_summary,
)
from docrt.paths import (
    ValidationError,
    ensure_output_not_locked,
    ensure_unlocked_for_read,
    validate_input_path,
    validate_output_path,
)
from docrt.read_ops import read_xlsx
from docrt.runtime_env import assert_mainline_runtime_for_path, confirmed_mainline_runtime


def patch_xlsx(
    input_path: str | Path,
    patch_path: str | Path,
    output_path: str | Path,
    *,
    dry_run: bool = False,
) -> dict[str, object]:
    source = validate_input_path(input_path, {".xlsx"})
    assert_mainline_runtime_for_path(source)
    patch_file = validate_input_path(patch_path, {".json"})
    target = validate_output_path(output_path)
    ensure_distinct_output(source, target)
    ensure_unlocked_for_read(source)
    ensure_unlocked_for_read(patch_file)
    if not dry_run:
        ensure_output_not_locked(target)
    patch = load_patch(patch_file, expected_document_type="xlsx")

    try:
        import openpyxl
    except Exception as exc:
        raise RuntimeError("openpyxl is unavailable") from exc

    workbook = openpyxl.load_workbook(str(source))
    changes: list[dict[str, object]] = []
    conflicts: list[dict[str, object]] = []
    try:
        for operation in patch["operations"]:
            op_type = operation.get("type")
            if op_type == "set_cell":
                operation_changes = [_xlsx_set_cell(workbook, operation, dry_run=dry_run)]
            elif op_type == "set_range_values":
                operation_changes = _xlsx_set_range_values(workbook, operation, dry_run=dry_run)
            elif op_type == "add_sheet":
                operation_changes = [_xlsx_add_sheet(workbook, operation, dry_run=dry_run)]
            elif op_type == "rename_sheet":
                operation_changes = [_xlsx_rename_sheet(workbook, operation, dry_run=dry_run)]
            else:
                raise ValidationError(
                    ErrorCode.VALIDATION_FAILED,
                    f"Unsupported XLSX patch op: {op_type}",
                )
            for change in operation_changes:
                if change.get("conflict"):
                    conflicts.append(change)
                changes.append(change)
        if not dry_run:
            shutil.copyfile(source, target)
            workbook.save(str(target))
    finally:
        workbook.close()

    verification = None
    if not dry_run:
        with confirmed_mainline_runtime():
            verification = read_xlsx(target)
    return {
        "input_path": str(source),
        "patch_path": str(patch_file),
        "output_path": str(target),
        "dry_run": dry_run,
        "patch_summary": patch_summary(changes, 0, conflicts),
        "verification": verification_summary(verification),
    }


def _xlsx_set_cell(workbook, operation: dict[str, Any], *, dry_run: bool) -> dict[str, object]:
    sheet = _worksheet(workbook, require_string(operation, "sheet"))
    cell = require_string(operation, "cell")
    value = operation.get("value")
    old_value = sheet[cell].value
    expected_value = operation.get("expected_value")
    if "expected_value" in operation and expected_value != old_value:
        raise ValidationError(
            ErrorCode.VALIDATION_FAILED,
            f"Cell {sheet.title}!{cell} expected_value mismatch",
        )
    if not dry_run:
        sheet[cell] = value
    return {
        "type": "set_cell",
        "sheet": sheet.title,
        "cell": cell,
        "old_value": old_value,
        "value": value,
        "planned_count": 1,
        "applied_count": 0 if dry_run else 1,
        "format_preservation": "preserve_existing_cell_style",
    }


def _xlsx_set_range_values(
    workbook, operation: dict[str, Any], *, dry_run: bool
) -> list[dict[str, object]]:
    sheet = _worksheet(workbook, require_string(operation, "sheet"))
    start_cell = require_string(operation, "start_cell")
    values = operation.get("values")
    if (
        not isinstance(values, list)
        or not values
        or not all(isinstance(row, list) for row in values)
    ):
        raise ValidationError(ErrorCode.VALIDATION_FAILED, "values must be a non-empty 2D list")
    expected_values = operation.get("expected_values")
    if expected_values is not None and not (
        isinstance(expected_values, list) and all(isinstance(row, list) for row in expected_values)
    ):
        raise ValidationError(ErrorCode.VALIDATION_FAILED, "expected_values must be a 2D list")
    start = sheet[start_cell]
    changes = []
    for row_offset, row_values in enumerate(values):
        for column_offset, value in enumerate(row_values):
            cell = sheet.cell(start.row + row_offset, start.column + column_offset)
            old_value = cell.value
            if expected_values is not None:
                try:
                    expected_value = expected_values[row_offset][column_offset]
                except IndexError as exc:
                    raise ValidationError(
                        ErrorCode.VALIDATION_FAILED,
                        "expected_values shape must match values shape",
                    ) from exc
                if expected_value != old_value:
                    raise ValidationError(
                        ErrorCode.VALIDATION_FAILED,
                        f"Cell {sheet.title}!{cell.coordinate} expected_values mismatch",
                    )
            if not dry_run:
                cell.value = value
            changes.append(
                {
                    "type": "set_cell",
                    "sheet": sheet.title,
                    "cell": cell.coordinate,
                    "old_value": old_value,
                    "value": value,
                    "planned_count": 1,
                    "applied_count": 0 if dry_run else 1,
                    "format_preservation": "preserve_existing_cell_style",
                }
            )
    return changes


def _xlsx_add_sheet(workbook, operation: dict[str, Any], *, dry_run: bool) -> dict[str, object]:
    name = require_string(operation, "name")
    if name in workbook.sheetnames:
        raise ValidationError(ErrorCode.VALIDATION_FAILED, f"Sheet already exists: {name}")
    if not dry_run:
        workbook.create_sheet(name)
    return {
        "type": "add_sheet",
        "name": name,
        "planned_count": 1,
        "applied_count": 0 if dry_run else 1,
    }


def _xlsx_rename_sheet(workbook, operation: dict[str, Any], *, dry_run: bool) -> dict[str, object]:
    old_name = require_string(operation, "old_name")
    new_name = require_string(operation, "new_name")
    if new_name in workbook.sheetnames:
        raise ValidationError(ErrorCode.VALIDATION_FAILED, f"Sheet already exists: {new_name}")
    sheet = _worksheet(workbook, old_name)
    if not dry_run:
        sheet.title = new_name
    return {
        "type": "rename_sheet",
        "old_name": old_name,
        "new_name": new_name,
        "planned_count": 1,
        "applied_count": 0 if dry_run else 1,
    }


def _worksheet(workbook, name: str):
    if name not in workbook.sheetnames:
        raise ValidationError(ErrorCode.VALIDATION_FAILED, f"Sheet not found: {name}")
    return workbook[name]
